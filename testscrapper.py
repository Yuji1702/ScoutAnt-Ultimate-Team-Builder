import requests
from bs4 import BeautifulSoup
import time
import random
import logging
from concurrent.futures import ThreadPoolExecutor
from typing import List, Dict
from openpyxl import Workbook, load_workbook
import os

logging.basicConfig(level=logging.INFO)

class ValorantDataCollector:
    """
    Collects Valorant player data from vlr.gg.
    """
    def __init__(self):
        """
        Initializes the ValorantDataCollector with the base URL and headers.
        """
        self.base_url = "https://www.vlr.gg"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.request_delay = (1.5, 3.0)  # Delay range for requests in seconds

    def _make_request(self, url: str) -> requests.Response:
        """
        Handles making a request to a URL with user-agent headers and delay.

        Args:
            url: The URL to request.

        Returns:
            The response object if successful, None otherwise.
        """
        time.sleep(random.uniform(*self.request_delay)) # Respectful delay
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()  # Raise HTTPError for bad responses (4xx or 5xx)
            return response
        except requests.exceptions.RequestException as e:
            logging.error(f"Request failed for {url}: {e}")
            return None

    def search_player(self, player_name: str) -> str:
        """
        Search for a player by name and return their profile URL.

        Args:
            player_name: The name of the player to search for.

        Returns:
            The player's profile URL if found, None otherwise.
        """
        search_url = f"{self.base_url}/search/?q={player_name.replace(' ', '%20')}&type=players"
        logging.info(f"Searching for player: {player_name}...")
        response = self._make_request(search_url)
        if not response:
            logging.error(f"Failed to retrieve search page for {player_name}")
            return None

        soup = BeautifulSoup(response.content, 'html.parser')
        player_link = soup.find('a', class_='search-item')

        if player_link:
            player_profile_url = self.base_url + player_link['href']
            logging.info(f"Player {player_name} found at: {player_profile_url}")
            return player_profile_url
        else:
            logging.warning(f"Player {player_name} not found.")
            return None

    def _extract_player_name(self, soup: BeautifulSoup) -> str:
        """
        Extracts the player name from the player profile page.

        Args:
            soup: BeautifulSoup object of the player profile page.

        Returns:
            The player name as a string, or 'Unknown' if not found.
        """
        name_tag = soup.find('h1', class_='wf-title')
        return name_tag.text.strip() if name_tag else 'Unknown'

    def _extract_stats_from_table(self, soup: BeautifulSoup) -> List[Dict]:
        """
        Extracts stats from the player's stats table on their profile page.

        Args:
            soup: BeautifulSoup object of the player profile page.

        Returns:
            A list of dictionaries, where each dictionary represents stats for an agent.
        """
        stats_table = soup.find('table', class_='wf-table')
        stats = []

        if not stats_table:
            logging.warning("Stats table not found on player page.")
            return stats  # Return empty list if no table

        try:
            stats_rows = stats_table.find('tbody').find_all('tr')
            for row in stats_rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    agent_img = cells[0].find('img')
                    agent_name = agent_img['alt'] if agent_img else 'Unknown'

                    def safe_number(cell):
                        if not cell or not cell.text.strip():
                            return 0.0  # Return 0.0 for empty cells

                        text_value = cell.text.strip()

                        # Check if the text contains a percentage and rounds played (for "Usage" column)
                        if '%' in text_value and '(' in text_value and ')' in text_value:
                            try:
                                # Extract percentage part (e.g., "50.0%")
                                percentage_str = text_value.split('%')[0] # Get part before '%'
                                return float(percentage_str) # Convert percentage to float
                            except ValueError:
                                logging.warning(f"Could not parse percentage from Usage: {text_value}")
                                return text_value # Return original text if percentage parsing fails

                        try:
                            # Try to convert to float for other numerical columns
                            return float(text_value)
                        except ValueError:
                            # If it's not a float, return the text as is (e.g., for "K:D" like "1.20")
                            return text_value


                    stats.append({
                        'Agent': agent_name,
                        'Usage': safe_number(cells[1]),
                        'Rounds Played': safe_number(cells[2]),
                        'Rating': safe_number(cells[3]),
                        'ACS': safe_number(cells[4]),
                        'K:D': safe_number(cells[5]),
                        'ADR': safe_number(cells[6]),
                        'KAST': safe_number(cells[7]),
                        'KPR': safe_number(cells[8]),
                        'APR': safe_number(cells[9]),
                        'First Kills Per Round': safe_number(cells[10]),
                        'First Deaths Per Round': safe_number(cells[11]),
                        'Kills': safe_number(cells[12]),
                        'Deaths': safe_number(cells[13]),
                        'Assists': safe_number(cells[14]),
                        'First Kills': safe_number(cells[15]),
                        'First Deaths': safe_number(cells[16])
                    })
            return stats

        except Exception as e:
            logging.error(f"Error processing stats table: {str(e)}")
            return [] # Return empty list in case of error

    def get_player_stats(self, url: str) -> Dict:
        """
        Extract stats for a single player, replacing missing values with 0.0.

        Args:
            url: The URL of the player's profile page.

        Returns:
            A dictionary containing the player's name and a list of their stats.
        """
        url_with_timespan = f"{url}?timespan=all"
        logging.info(f"Fetching stats from: {url_with_timespan}")
        response = self._make_request(url_with_timespan)
        if not response:
            logging.error(f"Failed to retrieve player stats page for {url}")
            return None

        soup = BeautifulSoup(response.content, 'html.parser')
        name = self._extract_player_name(soup)
        stats = self._extract_stats_from_table(soup)

        if stats:
            logging.info(f"Stats fetched successfully for player: {name}")
            return {'Name': name, 'Stats': stats}
        else:
            return None


def export_player_stats(player_data, filename):
    """Exports player stats to an Excel sheet."""
    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append([
            'Name', 'Agent', 'Usage', 'Rounds Played', 'Rating', 'ACS', 'K:D',
            'ADR', 'KAST', 'KPR', 'APR', 'First Kills Per Round', 'First Deaths Per Round',
            'Kills', 'Deaths', 'Assists', 'First Kills', 'First Deaths'
        ])

    for stat in player_data['Stats']:
        sheet.append([
            player_data['Name'],
            stat['Agent'],
            stat['Usage'],
            stat['Rounds Played'],
            stat['Rating'],
            stat['ACS'],
            stat['K:D'],
            stat['ADR'],
            stat['KAST'],
            stat['KPR'],
            stat['APR'],
            stat['First Kills Per Round'],
            stat['First Deaths Per Round'],
            stat['Kills'],
            stat['Deaths'],
            stat['Assists'],
            stat['First Kills'],
            stat['First Deaths']
        ])

    try:
        wb.save(filename)
        logging.info(f"Data exported to {filename}")
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")


def collect_data_for_player(player_name, collector, concurrent):
    """Collects data for a single player and saves it accordingly."""
    print(f"DEBUG: collect_data_for_player called for {player_name} with concurrent = {concurrent}") # DEBUG PRINT
    player_url = collector.search_player(player_name)
    if player_url:
        player_stats = collector.get_player_stats(player_url)
        if player_stats:
            filename = f"{player_name.replace(' ', '_')}_stats.xlsx" if concurrent else 'player_stats.xlsx'
            export_player_stats(player_stats, filename)

def collect_and_save_data(player_names: List[str], concurrent=False):
    """Collects data for multiple players and saves to separate Excel files if concurrent."""
    collector = ValorantDataCollector()

    logging.info(f"Collecting data for {len(player_names)} players...")

    if concurrent:
        logging.info("Running in concurrent mode... (Multiple Files)")
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(collect_data_for_player, player_name, collector, concurrent) for player_name in player_names]
            for future in futures:
                future.result()
    else:
        logging.info("Running in sequential mode... (Single File)")
        for player_name in player_names:
            collect_data_for_player(player_name, collector, concurrent)
            time.sleep(random.uniform(1.5, 3.0))

    logging.info("Data collection completed.")

if __name__ == "__main__":
    try:
        player_names = input("Enter the player names (comma-separated): ").split(',')
        player_names = [name.strip() for name in player_names]

        concurrent_input = input("Do you want to store files separately? (y/n): ").strip().lower()
        concurrent = concurrent_input == 'y'

        collect_and_save_data(player_names, concurrent=concurrent)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")